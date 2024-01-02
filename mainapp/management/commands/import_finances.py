import openpyxl
from django.core.management.base import BaseCommand
from mainapp.models import FieldEntry
from datetime import datetime, timedelta

class Command(BaseCommand):
    help = 'Import financial data from an Excel workbook'

    def handle(self, *args, **kwargs):
        # Prompt the user to clear existing data
        clear_data = input("Do you want to clear existing data before import? (y/n): ").lower()

        if clear_data == 'y':
            self.stdout.write("Clearing existing data...")
            FieldEntry.objects.all().delete()
            self.stdout.write("Existing data cleared.")

        workbook_name = 'auxiliaryScripts\my excel finances.xlsx'
        self.stdout.write(f'Importing data from {workbook_name}')

        # Load workbook
        workbook = openpyxl.load_workbook(workbook_name)

        # Sheet names to look for
        sheet_names = ['checking cibc', 'visa cibc', 'rbc finance']

        for sheet_name in sheet_names:
            if sheet_name in workbook.sheetnames:
                self.import_sheet(workbook[sheet_name])
            else:
                self.stdout.write(self.style.WARNING(f'Sheet "{sheet_name}" not found in workbook'))

        self.stdout.write(self.style.SUCCESS('Import completed successfully'))

    def import_sheet(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            date, description, withdrawals, deposits = row

            # Check for None and set to 0
            withdrawals = withdrawals if withdrawals is not None else 0
            deposits = deposits if deposits is not None else 0

            # Check if date is in serial format (Excel date format)
            if isinstance(date, int) or isinstance(date, float):
                date = datetime(1899, 12, 30) + timedelta(days=date)  # Convert to datetime

            transaction_type = 'deposit' if deposits else 'withdrawal'
            amount = abs(deposits or withdrawals)

            entry = FieldEntry.objects.create(
                date=date.date(),  # Extract the date part
                time=datetime.min.time(),  # Default time as midnight
                message=description,
                money=amount,
                type=transaction_type,
                read_status=False
            )

            self.stdout.write(f'Imported: {entry}')


